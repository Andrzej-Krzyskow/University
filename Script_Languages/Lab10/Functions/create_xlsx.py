import openpyxl


def create_xlsx(teachers_list):
    workbook = openpyxl.Workbook()

    sheet = workbook.active
    sheet.title = "Teachers"

    for row in range(1, len(teachers_list) + 1):
        sheet.cell(row=row, column=1, value=teachers_list[row - 1].fullname)
        sheet.cell(row=row, column=2, value=str(teachers_list[row - 1].email))

    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 35

    workbook.save("temp.xlsx")
