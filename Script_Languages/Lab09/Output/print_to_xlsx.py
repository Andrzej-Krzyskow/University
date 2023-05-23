import openpyxl

from Functions.avg_attack import avg_attack
from Functions.median_health import median_health
from Functions.count_costs import count_costs
from Functions.count_types import count_types
from Functions.total_cards import total_cards
from Functions.total_dust_cost import total_dust_cost


def print_to_xlsx(rows, filename):
    workbook = openpyxl.Workbook()

    sheet = workbook.active
    sheet.title = "Hearthstone cards"

    names = ["avg attack", "median health", "count types", "count costs", "total cards", "total dust cost"]
    values = [avg_attack(rows), median_health(rows), count_types(rows), count_costs(rows), total_cards(rows),
              total_dust_cost(rows)]

    for row in range(1, len(names)+1):
        sheet.cell(row=row, column=1, value=names[row-1])

    for row in range(1, len(names)+1):
        sheet.cell(row=row, column=2, value=str(values[row-1]))


    workbook.save(filename)
